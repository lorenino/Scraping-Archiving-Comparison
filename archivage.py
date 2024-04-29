import os
import requests
from bs4 import BeautifulSoup
from datetime import datetime
import difflib
from tqdm import tqdm
from openpyxl import load_workbook
from urllib.parse import urlparse
import smtplib
from email.mime.text import MIMEText
import logging
from bs4 import MarkupResemblesLocatorWarning, XMLParsedAsHTMLWarning
import warnings
import argparse

# Logging configuration
logging.basicConfig(filename='crawl.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Function to send an email with highlighted differences in HTML
def send_email(subject, diff_details, primary_url):
    sender_email = "your_email@gmail.com"
    receiver_email = "recipient_email@gmail.com"
    password = "your_password"

    # HTML content creation with differences formatting
    html_content = f"""
    <html>
    <body>
        <h2>{subject}</h2>
        <pre>{diff_details}</pre>
    </body>
    </html>
    """

    msg = MIMEText(html_content, 'html')
    msg['Subject'] = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Changes detected on {primary_url}"
    msg['From'] = sender_email
    msg['To'] = receiver_email

    try:
        server = smtplib.SMTP_SSL('smtp.example.com', 465)  # Example SMTP server
        server.login(sender_email, password)
        server.sendmail(sender_email, receiver_email, msg.as_string())
        server.quit()
        logging.info("Email sent successfully.")
    except Exception as e:
        logging.error("Error sending email: %s", e)

# Function to compare the two archives and send an email if differences are detected
def compare_archives(archive_folder, primary_url):
    try:
        archives = sorted([folder for folder in os.listdir(archive_folder) if os.path.isdir(os.path.join(archive_folder, folder))])
        if len(archives) < 2:
            return "At least two archives are required for comparison."
        latest_archive = os.path.join(archive_folder, archives[-1])
        latest_files = os.listdir(latest_archive)
        latest_content = {}
        for file in latest_files:
            with open(os.path.join(latest_archive, file), 'r', encoding="utf-8") as f:
                latest_content[file] = f.readlines()
        previous_archive = os.path.join(archive_folder, archives[-2])
        previous_files = os.listdir(previous_archive)
        previous_content = {}
        for file in previous_files:
            with open(os.path.join(previous_archive, file), 'r', encoding="utf-8") as f:
                previous_content[file] = f.readlines()
        differences = {}
        with tqdm(total=len(latest_files), desc="Comparing files", position=0, leave=True) as pbar:
            for file in latest_files:
                if file in previous_files:
                    latest_content_lines = latest_content[file]
                    previous_content_lines = previous_content[file]
                    diff = difflib.unified_diff(previous_content_lines, latest_content_lines, lineterm='')
                    differences[file] = list(diff)
                else:
                    differences[file] = ["The file does not exist in the previous archive."]
                pbar.update(1)
        for file, diff in differences.items():
            if diff and not all(line.startswith(' ') for line in diff):
                formatted_diff = '\n'.join([f"<span style='color:red;'>-{line}</span>" if line.startswith('-') else f"<span style='color:green;'>+{line}</span>" for line in diff])
                send_email("Changes Detected", f"In file: {file}<br>{formatted_diff}", primary_url)
    except Exception as e:
        logging.error("An error occurred while comparing archives: %s", e)

# Main function
def main():
    try:
        parser = argparse.ArgumentParser(description="Web crawler")
        parser.add_argument("file", help="Excel file containing URLs")
        args = parser.parse_args()

        urls = read_urls_from_excel(args.file)  # Read URLs from the Excel file
        if urls:
            primary_url = get_primary_url(urls)
            if primary_url:
                archive_folder = os.path.join("archives", primary_url)
                create_archive(primary_url, archive_folder, args)
                compare_archives(archive_folder, primary_url)  # Call function to compare the archives
            else:
                logging.warning("No valid URL found in the Excel file.")
    except Exception as e:
        logging.error("An error occurred while creating archives: %s", e)

# Function to read URLs from an Excel file
def read_urls_from_excel(filename):
    urls = []
    wb = load_workbook(filename)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        urls.append(row[0])  # Assuming URLs are in the first column
    return urls

# Function to extract HTML content from a URL and write it to a text file
def scrap_urls(urls, folder_name):
    try:
        with tqdm(total=len(urls), desc="Scraping pages", position=0, leave=True) as pbar:
            for url in urls:
                scrap_url(url, folder_name)
                pbar.update(1)
    except Exception as e:
        logging.error("An error occurred while scraping URLs: %s", e)
        return str(e)

# Function to extract HTML content from a URL and write it to a text file
def scrap_url(url, folder_name):
    try:
        response = requests.get(url)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            content = ""
            title_tag = soup.find('title')
            if title_tag:
                title = title_tag.get_text()
                content += f"Page Title: {title}\n\n"
            paragraphs = soup.find_all('p')
            for paragraph in paragraphs:
                content += f"{paragraph.get_text()}\n"
            parsed_url = urlparse(url)
            filename = parsed_url.netloc + parsed_url.path.replace("/", "_") + ".txt"
            filepath = os.path.join(folder_name, filename)
            with open(filepath, "w", encoding="utf-8") as file:
                file.write(content)
        else:
            logging.error("Request for %s failed with status code: %s", url, response.status_code)
    except Exception as e:
        logging.error("An error occurred while scraping URL %s : %s", url, e)

# Function to extract the primary URL from the list of URLs
def get_primary_url(urls):
    if urls:
        parsed_urls = [urlparse(url) for url in urls]
        domain_counts = {}
        for parsed_url in parsed_urls:
            domain = parsed_url.netloc  # Retrieve the full domain
            if domain in domain_counts:
                domain_counts[domain] += 1
            else:
                domain_counts[domain] = 1
        primary_domain = max(domain_counts, key=domain_counts.get)  # Retrieve the most frequent domain
        return primary_domain
    return None

# Function to create an archive for the primary URL
def create_archive(primary_url, archive_folder, args):
    try:
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        archive_name = os.path.join(archive_folder, f"{primary_url}_{timestamp}")
        os.makedirs(archive_name)
        urls = read_urls_from_excel(args.file)  # Read URLs from the Excel file
        scrap_urls(urls, archive_name)
        logging.info("Archive created for primary URL: %s", primary_url)
    except FileExistsError:
        logging.warning("Archive folder for primary URL already exists: %s", primary_url)
    except Exception as e:
        logging.error("An error occurred while creating archive for URL %s : %s", primary_url, e)

# Execute the main script
if __name__ == "__main__":
    # Ignore specific warnings
    warnings.filterwarnings("ignore", category=MarkupResemblesLocatorWarning)
    warnings.filterwarnings("ignore", category=XMLParsedAsHTMLWarning)

    main()

import os
import requests
from bs4 import BeautifulSoup
from datetime import datetime
import difflib
import shutil
from tqdm import tqdm
from openpyxl import load_workbook
from urllib.parse import urlparse
import smtplib
from email.mime.text import MIMEText
import schedule
import time
import warnings
import warnings
from bs4 import MarkupResemblesLocatorWarning, XMLParsedAsHTMLWarning


# Fonction pour envoyer un e-mail avec les différences mises en évidence en HTML
def send_email(subject, diff_details):
    sender_email = "mail@mail.mail"
    receiver_email = "mail@mail.mail"
    password = "Password"

    # Création du contenu HTML avec mise en forme des différences
    html_content = f"""
    <html>
    <body>
        <h2>{subject}</h2>
        <pre>{diff_details}</pre>
    </body>
    </html>
    """

    msg = MIMEText(html_content, 'html')
    msg['Subject'] = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {subject}"
    msg['From'] = sender_email
    msg['To'] = receiver_email

    try:
        server = smtplib.SMTP_SSL('ssl0.ovh.net', 465)  # Utilisation du serveur SMTP d'OVH
        server.login(sender_email, password)
        server.sendmail(sender_email, receiver_email, msg.as_string())
        server.quit()
        print("E-mail envoyé avec succès.")
    except Exception as e:
        print("Erreur lors de l'envoi de l'e-mail :", e)

# Fonction pour comparer les deux archives et envoyer un e-mail si des différences sont détectées
def compare_archives(archive_folder):
    try:
        archives = sorted([folder for folder in os.listdir(archive_folder) if os.path.isdir(os.path.join(archive_folder, folder))])
        if len(archives) < 2:
            return "Au moins deux archives sont nécessaires pour la comparaison."
        latest_archive = os.path.join(archive_folder, archives[-1])
        latest_files = os.listdir(latest_archive)
        latest_content = {}
        for file in latest_files:
            with open(os.path.join(latest_archive, file), 'r', encoding="utf-8") as f:
                latest_content[file] = f.readlines()
        previous_archive = os.path.join(archive_folder, archives[-2])
        previous_files = os.listdir(previous_archive)
        previous_content = {}
        for file in previous_files:
            with open(os.path.join(previous_archive, file), 'r', encoding="utf-8") as f:
                previous_content[file] = f.readlines()
        differences = {}
        with tqdm(total=len(latest_files), desc="Comparing files", position=0, leave=True) as pbar:
            for file in latest_files:
                if file in previous_files:
                    latest_content_lines = latest_content[file]
                    previous_content_lines = previous_content[file]
                    diff = difflib.unified_diff(previous_content_lines, latest_content_lines, lineterm='')
                    differences[file] = list(diff)
                else:
                    differences[file] = ["Le fichier n'existe pas dans l'archive précédente."]
                pbar.update(1)
        for file, diff in differences.items():
            if diff and not all(line.startswith(' ') for line in diff):
                formatted_diff = '\n'.join([f"<span style='color:red;'>-{line}</span>" if line.startswith('-') else f"<span style='color:green;'>+{line}</span>" for line in diff])
                send_email("[3s2i] Différences détectées", f"Dans le fichier : {file}<br>{formatted_diff}")
    except Exception as e:
        print("Une erreur s'est produite lors de la comparaison des archives : " + str(e))

# Fonction principale
def main():
    try:
        archive_folder = "archives"
        if not os.path.exists(archive_folder) or len(os.listdir(archive_folder)) < 2:
            print("Création de deux archives initiales...")
            for i in range(2):
                folder_name = os.path.join(archive_folder, "A_" + datetime.now().strftime("%Y-%m-%d_%H-%M-%S"))
                os.makedirs(folder_name)
                urls = read_urls_from_excel("urls_3s2i.xlsx")  # Lecture des URLs depuis le fichier Excel
                scrap_urls(urls, folder_name)
        if len(os.listdir(archive_folder)) >= 2:
            oldest_archive = min(os.listdir(archive_folder))
            shutil.rmtree(os.path.join(archive_folder, oldest_archive))
        folder_name = os.path.join(archive_folder, "A_" + datetime.now().strftime("%Y-%m-%d_%H-%M-%S"))
        os.makedirs(folder_name)
        urls = read_urls_from_excel("urls_3s2i.xlsx")  # Lecture des URLs depuis le fichier Excel
        scrap_urls(urls, folder_name)
        compare_archives(archive_folder)
    except Exception as e:
        print("Une erreur s'est produite :", str(e))

# Fonction pour lire les URLs depuis un fichier Excel
def read_urls_from_excel(filename):
    urls = []
    wb = load_workbook(filename)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        urls.append(row[0])  # Supposant que les URLs sont dans la première colonne
    return urls

# Fonction pour extraire le contenu HTML d'une URL et l'écrire dans un fichier texte
def scrap_urls(urls, folder_name):
    try:
        with tqdm(total=len(urls), desc="Scraping pages", position=0, leave=True) as pbar:
            for url in urls:
                scrap_url(url, folder_name)
                pbar.update(1)
    except Exception as e:
        return str(e)

# Fonction pour extraire le contenu HTML d'une URL et l'écrire dans un fichier texte
def scrap_url(url, folder_name):
    try:
        response = requests.get(url)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            content = ""
            title_tag = soup.find('title')
            if title_tag:
                title = title_tag.get_text()
                content += f"Titre de la page : {title}\n\n"
            paragraphs = soup.find_all('p')
            for paragraph in paragraphs:
                content += f"{paragraph.get_text()}\n"
            parsed_url = urlparse(url)
            filename = parsed_url.netloc + parsed_url.path.replace("/", "_") + ".txt"
            filepath = os.path.join(folder_name, filename)
            with open(filepath, "w", encoding="utf-8") as file:
                file.write(content)
        else:
            print(f"La requête pour {url} a échoué avec le code d'état: {response.status_code}")
    except Exception as e:
        print(str(e))

# Exécution du script principal
if __name__ == "__main__":
    # Ignorer les avertissements spécifiques
    warnings.filterwarnings("ignore", category=MarkupResemblesLocatorWarning)
    warnings.filterwarnings("ignore", category=XMLParsedAsHTMLWarning)

    # Planifier l'exécution de la fonction main toutes les minutes
    schedule.every(1).hours.do(main)
    
    # Boucle pour exécuter les tâches planifiées
    while True:
        schedule.run_pending()
        time.sleep(1)
